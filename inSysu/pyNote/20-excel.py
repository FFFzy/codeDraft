# pip install openpyxl
import openpyxl

# 新建表格并写入
wb = openpyxl.Workbook()
ws = wb.create_sheet(index=0)      
for i in range(1, 5):                    
    ws.cell(i, 1).value = "NAME" + str(i)   # 第i行第一列
    ws.cell(i, 2).value = "AGE" + str(i)    # 第i行第二列
    ws.cell(i, 3).value = "BIRTH" + str(i)  # 第i行第三列
wb.save("result.xlsx")

# 加载表格并读取
wb = openpyxl.load_workbook('./result.xlsx')
ws = wb.active 
for row in ws.rows:
    name = row[0].value
    age = row[1].value
    birth = row[2].value
    print(name, age, birth)